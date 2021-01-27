import os

import openpyxl
from openpyxl import Workbook
from case_auto_create import constants


class ExcelManager(object):
    """
    author: DuPanPan
    date  : 2021.01.18
    description: Excel表格管理
    """
    file_column = 25
    file_row = 0
    file_name = constants.OUTPUT_EXCEL_FILE_NAME
    sheet_name = constants.OUTPUT_EXCEL_SHEET_NAME
    output_dir = None
    file_path = None
    column_name_list = None

    @classmethod
    def set_file_column(cls, column_name_list: str) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 设置excel表格列名
        :param: column_name_list:excel表格列名表
        :return: 无返回值
        """
        cls.column_name_list = column_name_list

    @classmethod
    def set_file_column_list(cls, file_column: str) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 设置创建的excel的列数
        :param: file_column: excel的列数
        :return: 无返回值
        """
        cls.file_column = file_column

    @classmethod
    def set_file_name(cls) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 设置文件的名称
        :param: file_name: Excel文件名称
        :return: 无返回值
        """
        cls.file_name = input("Please excel name:")

    @classmethod
    def set_sheet_name(cls) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 设置表格也·2的名称
        :param: file_name: 创建的sheet页名称
        :return: 无返回值
        """
        cls.sheet_name = input("sheet name:")

    @classmethod
    def _file_create(cls, column_name_list) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 创建Excel文件
        :return: 无返回值
        """
        root_dir = os.path.dirname(os.path.abspath(__file__))
        cls.output_dir = os.path.join(root_dir, constants.OUTPUT_PRODUCT).replace("\\", "/")
        if not os.path.exists(cls.output_dir):
            try:
                os.makedirs(cls.output_dir)
            except Exception as e:
                print("Create dictionary %s failure" % cls.file_path)
                print(e)

        cls.file_path = os.path.join(cls.output_dir, cls.file_name + constants.EXCEL_FILE_SUFFIX).replace("\\", "/")
        print(cls.file_path)

        try:
            workbook = Workbook()
            # 创建带有index的sheet
            sheet = workbook.create_sheet(title=cls.sheet_name, index=0)
            # 将每列的title加入sheet中
            sheet.append(column_name_list)
            # 设置excel表的列宽
            sheet.sheet_format.defaultColWidth = constants.EXCEL_COL_WIDTH
            # 保存excel文件
            workbook.save(cls.file_path)
            print("make %s success" % cls.file_path)
        except Exception as e:
            print("create excel file failed. Error is %s" % e)
            exit(1)

    @classmethod
    def file_create(cls, column_name_list, case_list: dict) -> str:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 创建excel表格并填入数据
        :param: case_list，用例列表
        """
        # cls.set_file_name()
        # cls.set_sheet_name()
        cls._file_create(column_name_list)
        workbook = openpyxl.load_workbook(cls.file_path)
        worksheet = workbook[cls.sheet_name]
        for item in case_list:
            row_info = []
            for key in column_name_list:
                row_info.append(str(item.get(key)))
            worksheet.append(row_info)
        workbook.save(cls.file_path)
        print("write excel file %s success" % cls.file_path)
