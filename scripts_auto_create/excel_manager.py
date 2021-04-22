# -*- coding:utf-8 -*-
"""
author:      panpan.du
date:        2021.01.18
description: excel表格管理类
"""
import os
import openpyxl
from openpyxl import Workbook
from scripts_auto_create import constants


class ExcelManager(object):
    """
    author: DuPanPan
    date  : 2021.01.18
    description: Excel表格管理
    """
    file_column = 25
    file_path = constants.INPUT_EXCEL_PATH
    sheet_name = constants.INPUT_EXCEL_SHEET_NAME
    __case_script_list_path = None
    output_dir = None
    column_name_list = [constants.case_number, constants.case_script, constants.case_title,
                        constants.case_product, constants.case_module, constants.case_object,
                        constants.case_type, constants.case_classification, constants.case_check_point,
                        constants.case_card, constants.case_physical_env, constants.case_pre_condition,
                        constants.case_scene, constants.case_steps, constants.case_expect,
                        constants.case_priority, constants.case_is_smoke, constants.case_is_mini,
                        constants.case_is_auto, constants.case_time, constants.case_author, constants.case_notes,
                        constants.case_is_debug, constants.case_result, constants.case_is_upload]

    framework_excel_column_list = [constants.FRAMEWORK_EXCEL_CASE_ID, constants.FRAMEWORK_EXCEL_CASE_TITLE,
                                   constants.FRAMEWORK_EXCEL_CASE_VERSION, constants.FRAMEWORK_EXCEL_CASE_TYPE,
                                   constants.FRAMEWORK_EXCEL_CASE_LEVEL, constants.FRAMEWORK_EXCEL_CASE_LOCATION,
                                   constants.FRAMEWORK_EXCEL_CASE_BRANCH_ID,
                                   constants.FRAMEWORK_EXCEL_CASE_MODULE_ID, constants.FRAMEWORK_EXCEL_CASE_MODULE_NAME,
                                   constants.FRAMEWORK_EXCEL_CASE_STEP, constants.FRAMEWORK_EXCEL_CASE_RUN_ID,
                                   constants.FRAMEWORK_EXCEL_CASE_NEED_RUN]

    case_dict_list = list()
    title_index_dict = dict()



    @classmethod
    def case_title_get(cls, work_sheet: object):
        """
        author:DuPanPan
        date:2020.10.09
        description: 获取表头和对应索引号
        :param work_sheet:sheet表格
        """
        row_index_column_name = 1
        row_info = work_sheet[row_index_column_name]

        index = 0
        for cell in row_info:
            cls.title_index_dict[cell.value] = index
            index += 1

    @classmethod
    def __case_script_file_read(cls) -> None:
        """
        :author: DuPanPan
        :date  : 2020.03.02
        :description:读取用例表格
        :return: 无返回值
        """
        workbook = openpyxl.load_workbook(cls.file_path)
        worksheet = workbook[cls.sheet_name]
        cls.case_title_get(worksheet)
        # 从第2行读取数据到一个以脚本为关键字的字典中
        for row_info in worksheet.iter_rows(min_row=2):
            # 将表格的一行数据读取到一个字典
            case_info_dict = dict()
            for key in cls.column_name_list:
                case_info_dict[key] = row_info[cls.title_index_dict[key]].value
            cls.case_dict_list.append(case_info_dict)

    @classmethod
    def get_case_dict_list(cls) -> list:
        """
        :author: DuPanPan
        :date  : 2020.03.02
        :description:获取用例表格
        :return cls.case_dict_list: 无返回值
        """
        if not cls.case_dict_list:
            cls.__case_script_file_read()
        return cls.case_dict_list


    @classmethod
    def __case_script_file_create(cls) -> None:
        """
          author: DuPanPan
          date  : 2020.04.29
          description: 创建用例关系集文件
          :param: sheet_name: 创建的表名称
          :return: 无返回值
          """
        project_path = os.path.dirname(os.path.abspath(__file__))
        cls.__case_script_list_path = os.path.join(project_path, constants.CASE_SCRIPT_PART_PATH_NAME)
        try:
            workbook = Workbook()
            # 创建带有index的sheet
            sheet = workbook.create_sheet(title=cls.sheet_name, index=0)
            # 将每列的title加入sheet中
            sheet.append(cls.framework_excel_column_list)
            # 设置excel表的列宽
            sheet.sheet_format.defaultColWidth = 15
            # 保存excel文件
            workbook.save(cls.__case_script_list_path)
        except Exception as result:
            print("create excel file failed. Error is %s" % result)
            exit(1)

    @classmethod
    def create_framework_excel(cls, case_dict_List: list) -> None:
        """
        :author: DuPanPan
        :date  : 2020.03.02
        :description:创建framework使用的excel文件
        :return cls.case_dict_list: 无返回值
        """
        print(case_dict_List)
        cls.__case_script_file_create()
        print(cls.__case_script_list_path)
        workbook = openpyxl.load_workbook(cls.__case_script_list_path)
        worksheet = workbook[cls.sheet_name]
        for item in case_dict_List:
            row_info = []
            for key in cls.framework_excel_column_list:
                row_info.append(item.get(key, ""))
            worksheet.append(row_info)
        workbook.save(cls.__case_script_list_path)
        print("write excel file %s success" % cls.__case_script_list_path)
