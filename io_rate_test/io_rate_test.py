# -*- coding:utf-8 -*-
"""
author:      panpan.du
date:        2021.04.20
description: 测试io_rate峰值的自动化工具
"""

import subprocess
import os
import openpyxl
from openpyxl import Workbook
import logging  # 引入logging模块
import os.path
import time

# 第一步，创建一个logger
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # Log等级总开关
# 第二步，创建一个handler，用于写入日志文件
rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
log_path = os.path.dirname(os.getcwd()) + '/Logs/'
log_name = log_path + rq + '.log'
logfile = log_name
fh = logging.FileHandler(logfile, mode='w')
fh.setLevel(logging.INFO)  # 输出到file的log等级的开关
# 第三步，定义handler的输出格式
formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
fh.setFormatter(formatter)
# 第四步，将logger添加到handler里面
logger.addHandler(fh)

class ExcelManager(object):
    """
    author: DuPanPan
    date  : 2021.01.18
    description: Excel表格管理
    """
    file_name = "iorate和带宽统计表格"
    sheet_name = ""
    output_dir = None
    file_path = None
    column_name_list = None
    excel_column_list = ["thread", "io_rate", "bandwidth"]

    @classmethod
    def file_create(cls) -> None:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 创建Excel文件
        :return: 无返回值
        """
        root_dir = os.path.dirname(os.path.abspath(__file__))
        cls.output_dir = os.path.join(root_dir, "product").replace("\\", "/")
        if not os.path.exists(cls.output_dir):
            try:
                os.makedirs(cls.output_dir)
            except Exception as e:
                print("Create dictionary %s failure" % cls.file_path)
                print(e)

        cls.file_path = os.path.join(cls.output_dir, cls.file_name + "xlsx").replace("\\", "/")
        print(cls.file_path)


    @classmethod
    def create_sheet(cls, result_list: list, sheet_name) -> str:
        """
        author: DuPanPan
        date  : 2021.01.18
        description: 创建excel表格并填入数据 x
        :param case_list:用例列表
        """
        try:
            workbook = Workbook()
            # 创建带有index的sheet
            sheet = workbook.create_sheet(sheet_name, index=0)
            # 将每列的title加入sheet中
            sheet.append(cls.column_name_list)
            # 设置excel表的列宽
            sheet.sheet_format.defaultColWidth = 15
            # 保存excel文件
            workbook.save(cls.file_path)
            print("make %s success" % cls.file_path)
        except Exception as e:
            print("create excel file failed. Error is %s" % e)
            exit(1)

        workbook = openpyxl.load_workbook(cls.file_path)
        worksheet = workbook[sheet_name]
        for item in result_list:
            worksheet.append(item)
        workbook.save(cls.file_path)
        print("write excel file %s success" % cls.file_path)




class IoRateTest(object):
    # vdbench_xml
    vdb_xml_file_context = "data_errors=1\nsd=default,journal=/root/ps3test/basicio_vdbench/journal,openflags=directio," \
                           "thread={io_thread}\nsd=sd1,lun={lun}\nwd=default,xfersize={xsersize},rdpct={rdpct}," \
                           "seekpct={seekpct}\nwd=wd1,sd=sd1\nrd=rd1,wd=wd*,warmup=5,elapsed=120,interval=1," \
                           "openflags=directio,iorate=max"
    vdb_xml_file_name = "/root/ps3test/basicio_vdbench/vdb_xml"

    rdpct_list = ["100"]
    seekpct_list = ["50"]
    thread_list = [4]
    result_list = list()
    lun = None
    index = 0

    @classmethod
    def create_vd(cls):
        cmd = "cli /c0 add vd r5 drives=0:0-3"
        rtn, output = subprocess.getstatusoutput(cmd)
        if rtn == 0:
            logger.info("VD 创建成功")
        else:
            logger.info("VD 创建失败：\n%s" % output)
            exit(1)
        cmd = "lsscsi|grep SUGON|awk '{print $6}"
        rtn, output = subprocess.getstatusoutput(cmd)
        if rtn == 0:
            cls.lun = output
        if cls.lun is None or cls.lun == "":
            logger.info("无法获取盘符")

    @classmethod
    def init_vdb_file(cls, io_thread, xfersize, rdpct, seekpct):
        vdb_xml_file_context = cls.vdb_xml_file_context.format(io_thread=io_thread, lun=cls.lun,
                                                               xfersize=xfersize, rdpct=rdpct, seekpct=seekpct)
        logger.info(vdb_xml_file_context)
        cmd = "echo {} > {}".format(vdb_xml_file_context, cls.vdb_xml_file_name)
        rtn, output = subprocess.getstatusoutput(cmd)
        if rtn != 0:
            logger.info("写xml文件失败")
            exit(1)

        cmd = "/home/vdbench/vdbench -f {}|grep avg|awk 'print $3, $4'".format(cls.vdb_xml_file_name)
        rtn, output = subprocess.getstatusoutput(cmd)
        if rtn == 0:
            result = list()
            iorate = output.split(" ")[0]
            bandwidth = output.split(" ")[-1]
            result.append()
            logger.info("{}：当数据块为{}，rdpct={}, seekpct={},thread={}时，iorate为：{}，带宽为{}MB/sec".format(cls.index),
                        xfersize, rdpct, seekpct, iorate, bandwidth)
